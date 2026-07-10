import frappe


@frappe.whitelist()
def get_collections():
    collections = frappe.db.sql("""
        SELECT c.name, c.collection_name, c.finalized_on,
               COUNT(ci.name) AS product_count
        FROM `tabCollection` c
        LEFT JOIN `tabCollection Item` ci ON ci.parent = c.name
        GROUP BY c.name
        ORDER BY c.finalized_on DESC, c.creation DESC
    """, as_dict=True)
    return collections


@frappe.whitelist()
def get_collection_items(collection_name):
    items = frappe.db.sql("""
        SELECT
            ci.item,
            ci.item_name,
            ci.source_site,
            ci.product_url,
            i.image,
            i.description,
            ip.price_list_rate AS price,
            ip.currency
        FROM `tabCollection Item` ci
        LEFT JOIN `tabItem` i ON i.name = ci.item
        LEFT JOIN `tabItem Price` ip
            ON ip.item_code = ci.item AND ip.price_list = 'Competitor USD'
        WHERE ci.parent = %(collection_name)s
        ORDER BY ci.idx
    """, {"collection_name": collection_name}, as_dict=True)
    return items


def _fetch_image(url):
    if not url:
        return None, None
    try:
        import io, requests
        r = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code != 200:
            return None, None
        content_type = r.headers.get("content-type", "")
        buf = io.BytesIO(r.content)
        if "webp" in content_type:
            try:
                from PIL import Image
                img = Image.open(buf).convert("RGBA")
                out = io.BytesIO()
                img.save(out, format="PNG")
                out.seek(0)
                return out, "png"
            except Exception:
                return None, None
        ext = "png" if "png" in content_type else "gif" if "gif" in content_type else "jpeg"
        buf.seek(0)
        return buf, ext
    except Exception:
        return None, None


@frappe.whitelist()
def generate_factory_excel(collection_name):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    items = get_collection_items(collection_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Factory Sheet"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["#", "Style Number", "Product Name", "Description", "Image", "Product URL"]
    col_widths = [5, 22, 30, 50, 18, 40]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30

    IMG_ROW_HEIGHT = 90

    for idx, item in enumerate(items, 1):
        row = idx + 1
        values = [
            idx,
            item.get("item") or "",
            item.get("item_name") or "",
            item.get("description") or "",
            "",  # image cell — filled below
            item.get("product_url") or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = center if col == 1 else left
            cell.border = border
        ws.row_dimensions[row].height = IMG_ROW_HEIGHT

        img_url = item.get("image") or ""
        buf, ext = _fetch_image(img_url)
        if buf:
            try:
                xl_img = XLImage(buf)
                # Scale to fit row height (~90pt ≈ 120px at 96dpi)
                scale = 120 / max(xl_img.height, 1)
                xl_img.height = 120
                xl_img.width = int(xl_img.width * scale)
                col_letter = get_column_letter(5)
                ws.add_image(xl_img, f"{col_letter}{row}")
            except Exception:
                ws.cell(row=row, column=5, value=img_url)
        elif img_url:
            ws.cell(row=row, column=5, value=img_url)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    import base64
    encoded = base64.b64encode(buffer.read()).decode("utf-8")
    safe_name = (collection_name or "collection").replace(" ", "_").replace("/", "-")
    return {"filename": f"{safe_name}_factory.xlsx", "data": encoded}
